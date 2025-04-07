import openpyxl
import os
import random
import pandas as pd

def create_demo_csv(name):
    initial_wb = openpyxl.Workbook()
    initial_wb["Sheet"].title = "Sheet1"
    initial_wb["Sheet1"]
    initial_wb["Sheet1"]["A1"] = "Temp"
    initial_wb["Sheet1"]["B1"] = "Weight"
    temp_num = 35
    weight_num = 35
    weight_loss = -1 * float(5/145)
    for i in range(2,500-35+3): # was (35,501), starts at row2 because row1 is occupied by column names, 500 is the temp it ends at - 35 since we're starting at 35, plus 3 to offset how we are starting at row2
        initial_wb["Sheet1"].cell(row=i, column=1).value = temp_num
        initial_wb["Sheet1"].cell(row=i, column=2).value = weight_num # 35=35, 180=30, 500=?? 145-5 (5/145)=0.0344827   325-25 470-20 
        temp_num += 1 
        weight_num += weight_loss
    for i in range(2,500-35+3): # iterating through temps and slightly changing each to seem more natural
        random_num = random.random()/2 - .25
        initial_wb["Sheet1"].cell(row=i, column=1).value += random_num
    initial_wb.save("initial_wb.xlsx")
    # Converts the initial excel file to a csv file
    pd_from_xl = pd.read_excel("initial_wb.xlsx")
    pd_from_xl.to_csv(f"{name}.csv", index=False)
    os.unlink("initial_wb.xlsx") # deletes the initial excel file, not needed anymore

os.chdir("Thumb_Drive_Placeholder")
create_demo_csv("Sample001")
create_demo_csv("Sample002")
create_demo_csv("Sample003")
os.chdir("..")

# Creates demo master excel file
def create_demo_master_excel(name):
    master_wb = openpyxl.Workbook()
    master_wb["Sheet"].title = "Sample001"
    master_wb.create_sheet("Equation")
    master_wb["Equation"]["A1"] = "(a/b)+(x/y)/2=z" # Not the real equation for confidentiality
    master_wb.save(f"{name}.xlsx")

create_demo_master_excel("master_wb")
print("Demo files created")
### DEMO FILES CREATED ###