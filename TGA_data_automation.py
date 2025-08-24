import openpyxl
import os
import pandas as pd
import csv
import matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference
import glob
import create_demo_files
import clear_demo_files
import sys
import time
from io import BytesIO
from PIL import Image


def main():
    # Searches for all csv files in CWD
    csv_files = glob.glob(rf"*.csv")
    if csv_files:
        print("CSV files found")
    else:
        print("No CSV files found")
        sys.exit(1)
    # Iterates over each of the files and process them
    for csv_file in csv_files:
        rowlist = []
        temps = []
        weights = []
        target1 = float(180)
        target2 = float(500)
        with open(f"{csv_file}") as file:
            reader = csv.DictReader(file)
            for row in reader:
                temps.append(float(row['Temp']))
                weights.append(float(row['Weight']))
                rowlist.append(row)
        closest_dict1 = min(rowlist, key=lambda x: abs(float(x['Temp']) - target1))
        closest_dict2 = min(rowlist, key=lambda x: abs(float(x['Temp']) - target2))
        """
        The 2 target Temp/Weight value pairs have been found,
        now we will add them to the master>sample sheet,
        run the equation and add the result purity.
        """
        mb = openpyxl.load_workbook('master_wb.xlsx') 
        general_equation = mb['Equation']['A1'].value
        evaluated_equation = (
            general_equation
            .replace("a", str(closest_dict1['Temp']))
            .replace("b", str(closest_dict1['Weight']))
            .replace("x", str(closest_dict2['Temp']))
            .replace("y", str(closest_dict2['Weight']))
        )
        expression = evaluated_equation.split('=')[0] # Strip the "=z" part to isolate the expression
        # Use eval to calculate the value of the expression
        z = eval(expression)


        sheetname = csv_file.split('.')[0]
        if sheetname in mb.sheetnames:
            print("Existing sample, creating duplicate with extension")
            counter = 1
            new_sheetname = sheetname
            while new_sheetname in mb.sheetnames:
                new_sheetname = "%s_%03d" % (sheetname, counter)
                counter += 1
            sheetname = new_sheetname
        else: 
             print("New sample found")
        sample_sheet = mb.create_sheet(sheetname)
        sample_sheet['A1'] = "Temp"
        sample_sheet['B1'] = "Weight"
        sample_sheet['C1'] = "Purity"
        sample_sheet['A2'] = float(closest_dict1['Temp'])
        sample_sheet['B2'] = float(closest_dict1['Weight'])
        sample_sheet['A3'] = float(closest_dict2['Temp'])
        sample_sheet['B3'] = float(closest_dict2['Weight'])
        sample_sheet['C2'] = z
        # Adjust column widths to fit their contents
        for column in sample_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get column letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2  # Add padding
            sample_sheet.column_dimensions[column_letter].width = adjusted_width

        # Add graphs
        plt.figure(figsize=(10, 6))
        plt.plot(temps, weights, label=sheetname, color='blue', linewidth=2)
        plt.title(f'Weight loss curve for {sheetname}')
        plt.xlabel('Temperature (°C)')
        plt.ylabel('Weight (mg)')
        plt.grid(True)
        plt.legend()
        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()
        img = openpyxl.drawing.image.Image(buf)
        img.width = 640
        img.height = 320
        sample_sheet.add_image(img, 'E5')
        
        # Saves the master workbook
        mb.save("master_wb.xlsx")
        buf.close()
        print(rf"'{sample_sheet.title}' Data added to master excel file")
    print("Closing program...")
    time.sleep(10)

if __name__ == "__main__":
    main()
    print("Data entry complete.")
